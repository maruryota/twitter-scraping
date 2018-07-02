#!/usr/bin/env python
# -*- coding: utf-8 -*-

import json, config #標準のjsonモジュールとconfig.pyの読み込み
from requests_oauthlib import OAuth1Session #OAuthのライブラリの読み込み
import MeCab
from collections import Counter
import csv
import openpyxl as px
import re 

mecab = MeCab.Tagger("mecabrc")
mecab.parse("mecabrc")

wb = px.load_workbook('twitter-jpn.xlsx')

def ma_parse(sentence, filter="名詞"):
    node = mecab.parseToNode(sentence)
    while node:
        if node.feature.startswith(filter):
            yield node.surface
        node = node.next


def excel_input(names, att_names, tweet_all, counts, followers):

    # excelオープン

    sheetnames = wb.get_sheet_names()
    print(sheetnames)

    if len(names) == len(att_names) and len(tweet_all) == len(names):
        pass 
    else:
        print('error')
        exit()

    for index, att_name in enumerate(att_names):
        if att_name in sheetnames:
            sheet = wb[att_name]
            excel_data(tweet_all[index], names[index], att_name, sheet, counts[index], followers[index])
        else:
            sheet = wb.create_sheet(title=att_name)
            print("make " + att_name)
            excel_data(tweet_all[index], names[index], att_name, sheet, counts[index], followers[index])




def excel_data(tweets, name, att_name, sheet, count, followers):

    # ヘッダーの入力
    sheet.cell(row=1, column=1).value = name
    sheet.cell(row=1, column=2).value = att_name
    sheet.cell(row=1, column=3).value = "ツイート数" + str(count)
    sheet.cell(row=1, column=4).value = "フォロワー数" + str(followers)

    for i, c in enumerate('abcde', start=1): 
        sheet.column_dimensions[c].width = 27

    results = []

    # 解析の実行とカウント
    for tweet in tweets:
        text = str(tweet)
        words = [word for word in ma_parse(text)]
        results += words

    counter = Counter(results)
    r = 2

    for word, cnt in counter.most_common():
        sheet.cell(row=r, column=1).value = str(word)
        sheet.cell(row=r, column=2).value = int(cnt)
    
        r += 1

def clean_text(text):

    cleaned = text
    cleaned = dest_http(cleaned)
    cleaned = dest_att(cleaned)
    cleaned = dest_tag(cleaned)
    cleaned = dest_dust(cleaned)

    return cleaned

def dest_http(text):

    cleaned = re.sub(r"(https?|ftp)(:\/\/[-_\.!~*\'()a-zA-Z0-9;\/?:\@&=\+\$,%#]+)", '', text)
    return cleaned

def dest_att(text):

    cleaned = re.sub(r'\@.\w*', '', text)
    return cleaned

def dest_tag(text):
    
    cleaned = re.sub(r'#.\w*', '', text)
    return cleaned

def dest_dust(text):

    cleaned = re.sub(r'[@:/,.();&%-・\[\]一ー]', '', text)
    return cleaned


alphaReg = re.compile(r'^[a-zA-Z]+$')

def isalpha(s):
    return alphaReg.match(s) is not None

    




class pycolor:
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    PURPLE = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    END = '\033[0m'
    BOLD = '\038[1m'
    UNDERLINE = '\033[4m'
    INVISIBLE = '\033[08m'
    REVERCE = '\033[07m'

CK = config.CONSUMER_KEY
CS = config.CONSUMER_SECRET
AT = config.ACCESS_TOKEN
ATS = config.ACCESS_TOKEN_SECRET
twitter = OAuth1Session(CK, CS, AT, ATS) #認証処理

url = "https://api.twitter.com/1.1/statuses/user_timeline.json" #タイムライン取得エンドポイント

find_tweets = []
tweet_count = []
person_list = ['@bigstonebtc', '@1000crypto', '@TrendStream', '@bokujyuumai', '@junbhirano', '@sen_axis', '@Whiskey_bonbon_', '@nishinokazu', '@miner_taro', '@yanyanchan2020']
name_list = []
followers_count = []

for person in person_list:

    find_tweets_person = []
    params ={'screen_name' : person, 'count' : 500, 'tweet_mode' : 'extended'}
    res = twitter.get(url, params = params)
    timelines = None

    if res.status_code == 200: #正常通信出来た場合
        timelines = json.loads(res.text) #レスポンスからタイムラインリストを取得

        if timelines == None:
            print('not found timelines')
            exit()

        first = timelines[0]
        print(pycolor.GREEN + first['user']['name'] + pycolor.END)
        print('start ' + first['created_at'])

        name_list.append(first['user']['name'])
        followers_count.append(first['user']['followers_count'])
        tweets = []
        count = 0

        for line in timelines: #タイムラインリストをループ処理

            
            datetime = line['created_at'].split()

            if int(datetime[2]) >= 21:
                split_tweet = line["full_text"].split()
                tweets.append(split_tweet)
                count += 1
                if line == timelines[-1]:
                    print('This is twitter junky')

            elif int(datetime[2]) < 21:
                print('end ' + line['created_at'])
                break
        
        tweet_count.append(count)

        # ごめんなさい冗長なコード書いてます反省してます。
        for tweet in tweets:

            # 配列の先頭に＠が含まれていた場合、ほぼ返信ツイートであるため
            # 排除して処理の対象から除き、次のツイートへいく
            if '@' in tweet[0]:
                # print(pycolor.RED + ''.join(tweet) + pycolor.END)
                continue

            # かくツイートに対してループを回す
            for index, block in enumerate(tweet):
                if "http" in block:
                    pass
                elif isalpha(block[-1]) == True and index != len(block)-1 and len(block) != 1 and index+1 < len(tweet):
                    if isalpha(tweet[index+1][0]) == True:
                        conbination = block + tweet[index+1]
                        conbination = clean_text(str(conbination))
                        if conbination != "":
                            find_tweets_person.append(conbination)
                        del tweet[index+1]
                else:
                    # print(block)

                    sentence = clean_text(str(block))
                    if sentence != "":
                        find_tweets_person.append(sentence)
    
        find_tweets.append(find_tweets_person)

    else: #正常通信出来なかった場合
        print("Failed: %d" % res.status_code)


# excelの入力
excel_input(name_list, person_list, find_tweets, tweet_count, followers_count)


# 以下の処理は　https://a-zumi.net/python-ma-parse-noun/ を参考にしている
found = []

for personal in find_tweets:
    for tweet in personal:
        text = str(tweet)
        words = [word for word in ma_parse(text)]
        found += words

counter = Counter(found)
rows = []
total_tweet = 0
total_follower = 0

r = 2

for i in tweet_count:
    total_tweet += i 

for i in followers_count:
    total_follower += i

for word, cnt in counter.most_common():
    row = [str(word), str(cnt)]
    rows.append(row)

    sheet = wb['Sheet1']
    sheet.cell(row=1, column=1).value = 'Total'
    sheet.cell(row=1, column=3).value = "ツイート数  " + str(total_tweet)
    sheet.cell(row=1, column=4).value = "フォロワー数  " + str(total_follower)
    sheet.cell(row=r, column=1).value = str(word)
    sheet.cell(row=r, column=2).value = int(cnt)
    r += 1

with open('twitter_jpa.csv', 'w') as f:
    writer = csv.writer(f, lineterminator='\n')
    writer.writerows(rows)



wb.save('twitter-jpn.xlsx')    


